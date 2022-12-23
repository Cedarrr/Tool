# -*- coding: utf-8 -*-
"""
Created on Thu Dec  8 21:57:16 2022

@author: cedar
"""

from openpyxl import load_workbook
from openpyxl.styles import *
import xlwt

wb=load_workbook('test.xlsx')
for ws in wb.worksheets:

    # 获取工作表的有数据的范围
    data_range=ws.calculate_dimension()

    # 为有数据的部分设置边框
    for row in ws[data_range]:
        for cell in row:
            cell.border=Border(top=Side(style='dashed'),
                           bottom=Side(style='dashed'))
            cell.font=Font(name='微软雅黑')
            #cell.alignment=Alignment(horizontal='left')
    # 冻结
    #ws.freeze_panes='H5'
wb.save('test.xlsx')



